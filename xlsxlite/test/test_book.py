import os
import pytest
import random
import shutil
import string
import time

from mock import patch
from openpyxl.reader.excel import load_workbook
from unittest import skip
from xlsxlite.book import XLSXBook
from .base import XLSXTest


class BookTest(XLSXTest):

    def setUp(self):
        super().setUp()

        os.mkdir("_tests")

    def tearDown(self):
        super().tearDown()

        shutil.rmtree("_tests")

    def test_empty(self):
        book = XLSXBook()
        book.finalize(to_file="_tests/empty.xlsx")

        book = load_workbook(filename="_tests/empty.xlsx")
        assert len(book.worksheets) == 1
        assert book.worksheets[0].title == "Sheet1"

    def test_simple(self):
        book = XLSXBook()
        sheet1 = book.add_sheet("People")
        sheet1.append_row("Name", "Email")
        sheet1.append_row("Jim", "jim@acme.com")
        sheet1.append_row("Bob", "bob@acme.com")

        book.add_sheet("Empty")

        # insert a new sheet at a specific index
        book.add_sheet("New first", index=0)

        book.finalize(to_file="_tests/simple.xlsx")

        book = load_workbook(filename="_tests/simple.xlsx")
        assert len(book.worksheets) == 3

        sheet1, sheet2, sheet3 = book.worksheets
        assert sheet1.title == "New first"
        assert sheet2.title == "People"
        assert sheet3.title == "Empty"

        self.assertExcelSheet(sheet1, [()])
        self.assertExcelSheet(sheet2, [("Name", "Email"), ("Jim", "jim@acme.com"), ("Bob", "bob@acme.com")])
        self.assertExcelSheet(sheet3, [()])

    def test_escaping(self):
        book = XLSXBook()
        sheet1 = book.add_sheet("Test")
        sheet1.append_row("< & > \" ! =")
        book.finalize(to_file="_tests/escaped.xlsx")

        book = load_workbook(filename="_tests/escaped.xlsx")
        self.assertExcelSheet(book.worksheets[0], [("< & > \" ! =",)])

    def test_sheet_limits(self):
        book = XLSXBook()
        sheet1 = book.add_sheet("Sheet1")

        # try to add row with too many columns
        column = ['x'] * 20000
        with pytest.raises(ValueError):
            sheet1.append_row(*column)

        # try to add more rows than allowed
        with patch('xlsxlite.book.XLSXSheet.MAX_ROWS', 3):
            sheet1.append_row('x')
            sheet1.append_row('x')
            sheet1.append_row('x')

            with pytest.raises(ValueError):
                sheet1.append_row('x')

    @skip
    def test_performance(self):
        def random_val():
            return ''.join(random.choices(string.ascii_uppercase + string.digits, k=16))

        t0 = time.time()

        book = XLSXBook()
        sheet1 = book.add_sheet("Sheet1")
        book.add_sheet("Sheet2")

        t1 = time.time()
        print(f"Initialized workbook in {t1 - t0} seconds")

        num_rows = 1024 * 1024
        for r in range(num_rows):
            row = [random_val() for c in range(10)]
            sheet1.append_row(*row)

        t2 = time.time()
        print(f"Wrote {num_rows} rows in {t2 - t1} seconds")

        book.finalize(to_file="_tests/simple.xlsx")

        t3 = time.time()
        print(f"Finalized XLSX file in {t3 - t2} seconds")
