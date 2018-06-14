import pytest
import random
import string
import xlsxwriter

from mock import patch
from openpyxl import Workbook
from openpyxl.worksheet.write_only import WriteOnlyCell
from openpyxl.writer.etree_worksheet import etree_write_cell
from xlsxlite.book import XLSXBook
from .base import tests_dir  # noqa


NUM_ROWS = 1000
NUM_COLS = 10

# generate some random strings to use as cell values
DATA = [''.join(random.choices(string.ascii_uppercase + string.digits, k=16)) for d in range(1000)]


@pytest.mark.usefixtures("tests_dir")
def test_xlxslite():
    book = XLSXBook()
    sheet1 = book.add_sheet("Sheet1")

    for r in range(NUM_ROWS):
        row = [DATA[(r * c) % len(DATA)] for c in range(NUM_COLS)]

        sheet1.append_row(*row)

    book.finalize(to_file="_tests/test.xlsx")


@pytest.mark.usefixtures("tests_dir")
@patch('openpyxl.worksheet.write_only.write_cell')
def test_openpyxl_etree(mock_write_cell):
    mock_write_cell.side_effect = etree_write_cell

    book = Workbook(write_only=True)
    sheet1 = book.create_sheet("Sheet1")

    for r in range(NUM_ROWS):
        row = [DATA[(r * c) % len(DATA)] for c in range(NUM_COLS)]

        cells = [WriteOnlyCell(sheet1, value=v) for v in row]
        sheet1.append(cells)

    book.save("_tests/test.xlsx")


@pytest.mark.usefixtures("tests_dir")
def test_openpyxl_lxml():
    book = Workbook(write_only=True)
    sheet1 = book.create_sheet("Sheet1")

    for r in range(NUM_ROWS):
        row = [DATA[(r * c) % len(DATA)] for c in range(NUM_COLS)]

        cells = [WriteOnlyCell(sheet1, value=v) for v in row]
        sheet1.append(cells)

    book.save("_tests/test.xlsx")


@pytest.mark.usefixtures("tests_dir")
def test_xlsxwriter():
    book = xlsxwriter.Workbook("_tests/test.xlsx")
    sheet1 = book.add_worksheet()

    for r in range(NUM_ROWS):
        row = [DATA[(r * c) % len(DATA)] for c in range(NUM_COLS)]

        for c, val in enumerate(row):
            sheet1.write(r, c, val)

    book.close()
