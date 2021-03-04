from datetime import datetime, timedelta

import pytest
from openpyxl.reader.excel import load_workbook
from unittest.mock import patch
from xlsxlite.writer import XLSXBook

from .base import XLSXTest, tests_dir  # noqa


@pytest.mark.usefixtures("tests_dir")
class BookTest(XLSXTest):
    def test_empty(self):
        book = XLSXBook()
        book.finalize(to_file="_tests/empty.xlsx")

        book = load_workbook(filename="_tests/empty.xlsx")
        assert len(book.worksheets) == 1
        assert book.worksheets[0].title == "Sheet1"

    def test_simple(self):
        book = XLSXBook()
        sheet1 = book.add_sheet("People")
        sheet1.append_row("Name", "Email")
        sheet1.append_row("Jim", "jim@acme.com")
        sheet1.append_row("Bob", "bob@acme.com")

        book.add_sheet("Empty")

        # insert a new sheet at a specific index
        book.add_sheet("New first", index=0)

        book.finalize(to_file="_tests/simple.xlsx")

        book = load_workbook(filename="_tests/simple.xlsx")
        assert len(book.worksheets) == 3

        sheet1, sheet2, sheet3 = book.worksheets
        assert sheet1.title == "New first"
        assert sheet2.title == "People"
        assert sheet3.title == "Empty"

        self.assertExcelSheet(sheet1, [()])
        self.assertExcelSheet(sheet2, [("Name", "Email"), ("Jim", "jim@acme.com"), ("Bob", "bob@acme.com")])
        self.assertExcelSheet(sheet3, [()])

    def test_cell_types(self):
        d1 = datetime(2013, 1, 1, 12, 0, 0)

        book = XLSXBook()
        sheet1 = book.add_sheet("Test")
        sheet1.append_row("str", True, False, 3, 1.23, d1)

        # try to write a cell value with an unsupported type
        with pytest.raises(ValueError):
            sheet1.append_row(timedelta(days=1))

        book.finalize(to_file="_tests/types.xlsx")

        book = load_workbook(filename="_tests/types.xlsx")
        self.assertExcelSheet(book.worksheets[0], [("str", True, False, 3, 1.23, d1)])

    def test_escaping(self):
        book = XLSXBook()
        sheet1 = book.add_sheet("Test")
        sheet1.append_row('< & > " ! =')
        book.finalize(to_file="_tests/escaped.xlsx")

        book = load_workbook(filename="_tests/escaped.xlsx")
        self.assertExcelSheet(book.worksheets[0], [('< & > " ! =',)])

    def test_sheet_limits(self):
        book = XLSXBook()
        sheet1 = book.add_sheet("Sheet1")

        # try to add row with too many columns
        column = ["x"] * 20000
        with pytest.raises(ValueError):
            sheet1.append_row(*column)

        # try to add more rows than allowed
        with patch("xlsxlite.writer.XLSXSheet.MAX_ROWS", 3):
            sheet1.append_row("x")
            sheet1.append_row("x")
            sheet1.append_row("x")

            with pytest.raises(ValueError):
                sheet1.append_row("x")
